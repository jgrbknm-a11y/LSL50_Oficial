#!/usr/bin/env python3
import argparse
import json
import re
import shutil
import sqlite3
import unicodedata
from datetime import datetime
from pathlib import Path

import pdfplumber
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
DB = ROOT / "data" / "lsl50_local.sqlite"
UPLOAD_LOGOS = ROOT / "public" / "uploads" / "logos"
PREVIEW_JSON = ROOT / "data" / "official_roster_import_preview.json"

SOURCE_ROOT = Path("/Users/joseramirez/Library/Mobile Documents/com~apple~CloudDocs/LSL50 primavera 2026")
ROSTER_ROOT = SOURCE_ROOT / "Roster Oficial Equipos"
LOGO_ROOT = SOURCE_ROOT / "Logos Equipos"

XLSX_SOURCES = [
    ("Sharks", ROSTER_ROOT / "Roster_Oficial_LSL50_Shark.xlsx", 17, "dayfirst"),
    ("Caribeños", ROSTER_ROOT / "Roster_Oficial_LSL50_Caribeños ENERO 2026.xlsx", 10, "iso"),
]

LOGO_SOURCES = {
    "league_logo_url": (LOGO_ROOT / "Logo oficial LSL50.PNG", "lsl50_official.png"),
    "bucaneros": (LOGO_ROOT / "Logo Bucaneros .PNG", "bucaneros.png"),
    "caribenos": (LOGO_ROOT / "Logo Caribeños .PNG", "caribenos.png"),
    "cerveceros": (LOGO_ROOT / "Cerveceros.jpg", "cerveceros.jpg"),
    "cubs": (LOGO_ROOT / "Logo Cubs.PNG", "cubs.png"),
    "hispanos": (LOGO_ROOT / "Logo Hispanos.JPG", "hispanos.jpg"),
    "sharks": (LOGO_ROOT / "Tiburones .PNG", "sharks.png"),
}

BUCANEROS = [
    (1, "Isaac", "Carciente", "09.15.1975"),
    (2, "Ender", "Davila", "07.08.1980"),
    (3, "Carlos", "Paez", ""),
    (4, "Angel", "Martinez", ""),
    (5, "Lionel", "Martinez", ""),
    (6, "Jose", "Sojo", ""),
    (7, "Freddy", "Alayon", ""),
    (8, "Dennis", "Quillarque", ""),
    (9, "Jose", "Gonzalez", "05.15.1985"),
    (10, "Alfonso", "Davila", ""),
    (11, "Gene", "Brandt", ""),
    (12, "Roger", "Ramirez", "11.20.1981"),
    (13, "Michael", "Gonzalez", "10.26.1979"),
    (14, "Kelwin", "Aristigueta", ""),
    (15, "Hugo", "Ortega", ""),
    (16, "Rafael", "Gimenez", ""),
    (17, "Douglas", "Gonzalez", "09.23.1979"),
    (18, "Atahualpa", "Soteldo", ""),
]

HISPANOS = [
    (51, "Orlando", "Hernandez", "12 MAYO 1968"),
    (21, "Victor", "Rymer", "17 FEBRERO 1961"),
    (13, "Jose", "Reyes", "13 SEPTIEMBRE 1961"),
    (23, "Carlos", "Soledispa", "23 FEBRERO 1980"),
    (11, "Ranier", "Cardenas", "10 OCTUBRE 1980"),
    (35, "Nestor", "Matheus", "9 SEPTIEMBRE 1968"),
    (10, "Eduard", "Peralta", "10 JULIO 1968"),
    (69, "Lenin", "Palacios", "5 ENERO 1969"),
    (1, "Rene", "Valdes", "18 SEPTIEMBRE 1983"),
    (25, "Brian", "Field", "23 SEPTIEMBRE 1972"),
    (3, "Rafael", "Munoz", "10 MARZO 1980"),
    (30, "Alex", "Almeida", "9 FEBRERO 1974"),
    (24, "Joel", "Alfaro", "19 MAYO 1977"),
    (4, "David", "Batista", "5 AGOSTO 1969"),
    (23, "Kelvin", "Joga", "23 FEBRERO 1974"),
    (16, "Avelo", "Banez", "11 ENERO 1978"),
    (12, "Osmanis", "Verdecia", "16 ENERO 1975"),
    (7, "Yordin", "Rodriguez", "27 ABRIL 1982"),
]

SPANISH_MONTHS = {
    "ENERO": 1,
    "FEBRERO": 2,
    "MARZO": 3,
    "ABRIL": 4,
    "MAYO": 5,
    "JUNIO": 6,
    "JULIO": 7,
    "AGOSTO": 8,
    "SEPTIEMBRE": 9,
    "SETIEMBRE": 9,
    "OCTUBRE": 10,
    "NOVIEMBRE": 11,
    "DICIEMBRE": 12,
}


def strip_accents(value: str) -> str:
    return "".join(c for c in unicodedata.normalize("NFKD", value) if not unicodedata.combining(c))


def slugify(value: str) -> str:
    value = strip_accents(value).lower()
    return re.sub(r"[^a-z0-9]+", "-", value).strip("-")


def clean(value) -> str:
    return str(value or "").strip()


def parse_date(value, mode: str) -> str | None:
    if not value:
        return None
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    text = clean(value)
    if not text:
        return None
    text = text.replace(".", "/").replace("-", "/")
    if mode == "spanish":
        parts = strip_accents(text.upper()).split()
        if len(parts) == 3 and parts[1] in SPANISH_MONTHS:
            return f"{int(parts[2]):04d}-{SPANISH_MONTHS[parts[1]]:02d}-{int(parts[0]):02d}"
    parts = [p for p in re.split(r"[/-]", text) if p]
    if len(parts) == 3 and all(p.isdigit() for p in parts):
        a, b, c = [int(p) for p in parts]
        if mode == "dayfirst":
            day, month, year = a, b, c
        else:
            month, day, year = a, b, c
        if year < 100:
            year += 1900
        return f"{year:04d}-{month:02d}-{day:02d}"
    return None


def age_on_today(birth: str | None) -> int | None:
    if not birth:
        return None
    born = datetime.strptime(birth, "%Y-%m-%d").date()
    today = datetime.today().date()
    return today.year - born.year - ((today.month, today.day) < (born.month, born.day))


def player(team, number, first, last, birth, source):
    birth_date = birth if birth and re.match(r"^\d{4}-\d{2}-\d{2}$", birth) else None
    return {
        "team": team,
        "number": clean(number),
        "first_name": clean(first).title(),
        "last_name": clean(last).title(),
        "birth_date": birth_date,
        "position": "",
        "source": source,
        "age": age_on_today(birth_date),
    }


def extract_xlsx(team: str, path: Path, header_row: int, date_mode: str):
    ws = load_workbook(path, data_only=True).active
    rows = []
    for r in range(header_row + 1, ws.max_row + 1):
        number = ws.cell(r, 1).value
        first = ws.cell(r, 2).value
        last = ws.cell(r, 3).value
        birth = ws.cell(r, 4).value
        if not first and not last:
            continue
        rows.append(player(team, number, first, last, parse_date(birth, date_mode), path.name))
    return rows


def extract_cubs_pdf():
    path = ROSTER_ROOT / "Roster_Oficial_LSL50_Cubs.pdf"
    rows = []
    with pdfplumber.open(path) as pdf:
        table = pdf.pages[0].extract_tables()[0]
    for raw in table[1:]:
        number, first, last, birth = raw[:4]
        if not first and not last:
            continue
        rows.append(player("Cubs", number, first, last, parse_date(birth, "monthfirst"), path.name))
    return rows


def manual_rows(team: str, rows, date_mode: str, source: str):
    return [player(team, number, first, last, parse_date(birth, date_mode), source) for number, first, last, birth in rows]


def build_roster():
    rows = []
    for team, path, header_row, date_mode in XLSX_SOURCES:
        rows.extend(extract_xlsx(team, path, header_row, date_mode))
    rows.extend(extract_cubs_pdf())
    rows.extend(manual_rows("Bucaneros", BUCANEROS, "monthfirst", "PHOTO-2026-01-29-10-45-34.jpg"))
    rows.extend(manual_rows("Hispanos", HISPANOS, "spanish", "Roster Oficial Hispanos .jpg"))

    seen = set()
    unique = []
    skipped = []
    for row in rows:
        key = (slugify(row["team"]), slugify(row["first_name"]), slugify(row["last_name"]), row["birth_date"] or "")
        if key in seen:
            skipped.append({**row, "reason": "duplicate name/date in same team"})
            continue
        seen.add(key)
        unique.append(row)
    return unique, skipped


def summarize(rows, skipped):
    by_team = {}
    missing_birth = []
    under_50 = []
    for row in rows:
        by_team[row["team"]] = by_team.get(row["team"], 0) + 1
        if not row["birth_date"]:
            missing_birth.append(row)
        elif row["age"] is not None and row["age"] < 50:
            under_50.append(row)
    return {
        "total_importable": len(rows),
        "by_team": dict(sorted(by_team.items())),
        "skipped_duplicates": skipped,
        "missing_birth_dates": missing_birth,
        "under_50": under_50,
    }


def connect():
    conn = sqlite3.connect(DB)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys=ON")
    return conn


def copy_logos(conn):
    UPLOAD_LOGOS.mkdir(parents=True, exist_ok=True)
    copied = {}
    for key, (src, out_name) in LOGO_SOURCES.items():
        if not src.exists():
            copied[key] = None
            continue
        dest = UPLOAD_LOGOS / out_name
        shutil.copy2(src, dest)
        url = "/public/uploads/logos/" + out_name
        copied[key] = url
        if key == "league_logo_url":
            conn.execute(
                "INSERT INTO app_settings (key, value, updated_at) VALUES (?, ?, CURRENT_TIMESTAMP) "
                "ON CONFLICT(key) DO UPDATE SET value=excluded.value, updated_at=CURRENT_TIMESTAMP",
                ("league_logo_url", url),
            )
        else:
            conn.execute("UPDATE teams SET logo_url=? WHERE slug=?", (url, key))
    return copied


def apply_import(rows):
    conn = connect()
    target_teams = sorted({r["team"] for r in rows})
    team_ids = {}
    for team in target_teams:
        slug = slugify(team)
        row = conn.execute("SELECT id FROM teams WHERE slug=? OR name=?", (slug, team)).fetchone()
        if not row:
            raise RuntimeError(f"Missing team in database: {team}")
        team_ids[team] = int(row["id"])

    with conn:
        for team, team_id in team_ids.items():
            conn.execute("DELETE FROM players WHERE team_id=?", (team_id,))
        for row in rows:
            cur = conn.execute(
                "INSERT INTO players (team_id, first_name, last_name, birth_date, number, position) VALUES (?, ?, ?, ?, ?, ?)",
                (team_ids[row["team"]], row["first_name"], row["last_name"], row["birth_date"], row["number"], row["position"]),
            )
            conn.execute("INSERT OR IGNORE INTO player_stats (player_id) VALUES (?)", (cur.lastrowid,))
        copied = copy_logos(conn)
    conn.close()
    return copied


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--apply", action="store_true", help="Import into the local SQLite database")
    args = parser.parse_args()

    rows, skipped = build_roster()
    summary = summarize(rows, skipped)
    payload = {"summary": summary, "players": rows}
    PREVIEW_JSON.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")

    print(json.dumps(summary, ensure_ascii=False, indent=2))
    print(f"Preview written to {PREVIEW_JSON}")
    if args.apply:
        copied = apply_import(rows)
        print(json.dumps({"applied": True, "logos": copied}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
